"""
文档处理管道优化系统
版本: 1.2
作者: 智能助手
"""

import os
import re
import hashlib
import logging
import concurrent.futures
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from functools import lru_cache
from datetime import datetime
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from pydub import AudioSegment
import moviepy.editor as mp
from bs4 import BeautifulSoup
import nltk
from nltk.corpus import stopwords
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
import numpy as np

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("document_processor.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# 初始化NLTK
nltk.download('punkt')
nltk.download('stopwords')

class DocumentProcessor:
    """文档处理核心类"""
    
    def __init__(self, max_workers: int = 4):
        self.max_workers = max_workers
        self.cache_dir = ".document_cache"
        os.makedirs(self.cache_dir, exist_ok=True)
        
        # 初始化质量评估模型
        self.quality_model = self._init_quality_model()
        
        # 注册文件处理器
        self.file_handlers = {
            '.pdf': self._process_pdf,
            '.docx': self._process_docx,
            '.xlsx': self._process_excel,
            '.csv': self._process_csv,
            '.mp3': self._process_audio,
            '.mp4': self._process_video,
            '.txt': self._process_text
        }

    # ========== 公开接口 ==========
    def process_document(self, file_path: str) -> Dict:
        """处理单个文档"""
        try:
            file_hash = self._get_file_hash(file_path)
            if self._check_cache(file_hash):
                return self._load_from_cache(file_hash)

            content, metadata = self._parse_file(file_path)
            chunks = self._chunk_content(content)
            processed = self._process_chunks(chunks)
            
            result = {
                "metadata": self._extract_metadata(content),
                "content": processed,
                "quality_score": self._assess_quality(processed),
                "file_hash": file_hash
            }
            
            self._save_to_cache(file_hash, result)
            return result
            
        except Exception as e:
            logger.error(f"处理文档失败: {str(e)}")
            raise DocumentProcessingError(f"文档处理失败: {str(e)}")

    def process_batch(self, file_paths: List[str]) -> Dict[str, Dict]:
        """批量处理文档"""
        results = {}
        with concurrent.futures.ProcessPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_path = {
                executor.submit(self.process_document, path): path
                for path in file_paths
            }
            
            for future in concurrent.futures.as_completed(future_to_path):
                path = future_to_path[future]
                try:
                    results[path] = future.result()
                except Exception as e:
                    results[path] = {"error": str(e)}
        return results

    # ========== 文件解析模块 ==========
    def _parse_file(self, file_path: str) -> Tuple[str, Dict]:
        """解析各种文件格式"""
        ext = os.path.splitext(file_path).lower()
        handler = self.file_handlers.get(ext)
        
        if not handler:
            raise UnsupportedFormatError(f"不支持的格式: {ext}")
            
        return handler(file_path)

    def _process_pdf(self, path: str) -> Tuple[str, Dict]:
        """处理PDF文件"""
        content = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                content.append(page.extract_text())
        return "\n".join(content), {"pages": len(pdf.pages)}

    def _process_excel(self, path: str) -> Tuple[str, Dict]:
        """处理Excel文件"""
        wb = load_workbook(path)
        content = []
        for sheet in wb:
            df = pd.DataFrame(sheet.values)
            content.append(f"Sheet: {sheet.title}\n{df.to_markdown()}")
        return "\n\n".join(content), {"sheets": len(wb.sheetnames)}

    def _process_audio(self, path: str) -> Tuple[str, Dict]:
        """处理音频文件"""
        audio = AudioSegment.from_file(path)
        # 此处应集成语音识别API
        return "[音频转写内容]", {"duration": len(audio)/1000}

    def _process_video(self, path: str) -> Tuple[str, Dict]:
        """处理视频文件"""
        video = mp.VideoFileClip(path)
        # 此处应集成字幕提取
        return "[视频字幕内容]", {"duration": video.duration}

    # ========== 文本分块优化 ==========
    def _chunk_content(self, text: str, max_length: int = 1000) -> List[str]:
        """优化后的分块算法"""
        # 基于句子边界的分块
        sentences = nltk.sent_tokenize(text)
        chunks = []
        current_chunk = []
        current_length = 0
        
        for sent in sentences:
            sent_length = len(sent)
            if current_length + sent_length > max_length:
                chunks.append(" ".join(current_chunk))
                current_chunk = []
                current_length = 0
                
            current_chunk.append(sent)
            current_length += sent_length
            
        if current_chunk:
            chunks.append(" ".join(current_chunk))
            
        return chunks

    # ========== 质量评估模块 ==========
    def _init_quality_model(self) -> LogisticRegression:
        """初始化质量评估模型"""
        # 示例训练数据
        texts = [
            ("高质量文本", 1),
            ("低质量内容 blah blah", 0),
            # 更多训练数据...
        ]
        vectorizer = TfidfVectorizer()
        X = vectorizer.fit_transform([t for t in texts])
        y = [t for t in texts]
        
        model = LogisticRegression()
        model.fit(X, y)
        return model

    def _assess_quality(self, content: List[str]) -> float:
        """评估文档质量"""
        features = self._extract_quality_features(" ".join(content))
        return self.quality_model.predict_proba([features])

    def _extract_quality_features(self, text: str) -> List[float]:
        """提取质量特征"""
        # 示例特征：长度、停用词比例、重复率
        words = nltk.word_tokenize(text)
        stop_words = set(stopwords.words('english'))
        
        return [
            len(text),
            len([w for w in words if w.lower() in stop_words]) / len(words),
            self._calculate_repetition_score(text)
        ]

    # ========== 元数据提取 ==========
    def _extract_metadata(self, text: str) -> Dict:
        """提取文档元数据"""
        metadata = {
            "title": self._extract_title(text),
            "author": None,
            "date": self._extract_dates(text)
        }
        
        # 从HTML/Markdown提取额外元数据
        if "<html>" in text.lower():
            soup = BeautifulSoup(text, 'html.parser')
            if soup.title:
                metadata["title"] = soup.title.string
                
        return metadata

    def _extract_title(self, text: str) -> Optional[str]:
        """提取标题"""
        lines = text.split("\n")
        for line in lines:
            stripped = line.strip()
            if stripped and len(stripped) < 100:
                return stripped
        return None

    # ========== 缓存和增量处理 ==========
    def _get_file_hash(self, path: str) -> str:
        """计算文件哈希"""
        with open(path, 'rb') as f:
            return hashlib.md5(f.read()).hexdigest()

    def _check_cache(self, file_hash: str) -> bool:
        """检查缓存是否存在"""
        return os.path.exists(self._cache_path(file_hash))

    def _save_to_cache(self, file_hash: str, data: Dict):
        """保存处理结果到缓存"""
        with open(self._cache_path(file_hash), 'w') as f:
            json.dump(data, f)

    def _cache_path(self, file_hash: str) -> str:
        """生成缓存路径"""
        return os.path.join(self.cache_dir, f"{file_hash}.json")

    # ========== 辅助方法 ==========
    @staticmethod
    def _calculate_repetition_score(text: str) -> float:
        """计算文本重复率"""
        words = nltk.word_tokenize(text)
        unique = set(words)
        return 1 - len(unique) / len(words)

    @staticmethod
    def _extract_dates(text: str) -> List[str]:
        """提取日期信息"""
        date_pattern = r"\d{4}-\d{2}-\d{2}"
        return re.findall(date_pattern, text)

# ========== 异常处理 ==========
class DocumentProcessingError(Exception):
    pass

class UnsupportedFormatError(DocumentProcessingError):
    pass

# ========== 单元测试 ==========
import unittest
from unittest.mock import patch

class TestDocumentProcessor(unittest.TestCase):
    def setUp(self):
        self.processor = DocumentProcessor()
        
    def test_pdf_processing(self):
        test_path = "sample.pdf"
        result = self.processor.process_document(test_path)
        self.assertIn("metadata", result)
        
    @patch.object(DocumentProcessor, '_process_pdf')
    def test_cache_mechanism(self, mock_process):
        test_path = "test.pdf"
        mock_process.return_value = ("test content", {})
        
        # 首次处理
        result1 = self.processor.process_document(test_path)
        # 第二次应使用缓存
        result2 = self.processor.process_document(test_path)
        self.assertEqual(result1["file_hash"], result2["file_hash"])
        mock_process.assert_called_once()

if __name__ == "__main__":
    # 示例用法
    processor = DocumentProcessor()
    results = processor.process_document("example.docx")
    print(f"处理结果: {json.dumps(results, indent=2)}")
    
    # 运行测试
    unittest.main()
